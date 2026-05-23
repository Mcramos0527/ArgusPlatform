# app/services/reconciler.py
# Wave 2 — Reconciliation engine.
#
# Crosses bank transactions (Step 1) against ERP Coliseo records
# (COBROS + PAGOS files) and produces a reconciliation report.
#
# Match logic:
#   Key = (amount, bank_hint) with date tolerance
#   For Mercado Pago: ±1 day tolerance (MP doesn't allow same-day download)
#   For all other banks: exact date or ±1 day tolerance
#
# Result statuses:
#   CONCILIADO       — found in both bank and ERP
#   PENDIENTE BANCO  — ERP has it, bank doesn't yet
#   PENDIENTE ERP    — bank has it, ERP doesn't have it

import logging
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger("argus.reconciler")

# Bank name hints from ERP "Detalle Valor" column
BANK_HINTS = {
    "ICBC":         ["ICBC"],
    "BBVA":         ["BBVA", "BANCO FRANCES", "FRANCES"],
    "Mercado Pago": ["MERCADO PAGO", "MP VENTAS", "MP GERENCIA", "MP "],
    "Bancor":       ["BANCOR"],
    "Nación":       ["NACION", "NACIÓN", "BNA"],
    "Galicia":      ["GALICIA"],
    "Cresium":      ["CRESIUM"],
}

# Banks that require ±1 day date tolerance
MP_BANKS = {"MP fondo azul", "MP fondo blanco",
            "Mercado Pago Gerencia (Fondo Azul)",
            "Mercado Pago Ventas (Fondo Blanco)"}


# ── Data models ───────────────────────────────────────────────────────────────

@dataclass
class ERPRecord:
    """A single record from COBROS or PAGOS ERP export."""
    fecha: Optional[date]
    comprobante: str
    nombre: str           # client or supplier name
    total: float          # total invoice amount
    medio_pago: str       # payment method description
    detalle_valor: str    # bank hint (e.g. "ICBC Dario Delfabro SRL")
    importe_valor: float  # actual amount transacted (may differ from total)
    fecha_valor: Optional[date]
    tipo: str             # "COBRO" or "PAGO"
    matched: bool = False


@dataclass
class ReconciliationLine:
    """Output line of the reconciliation report."""
    estado: str               # CONCILIADO / PENDIENTE BANCO / PENDIENTE ERP
    tipo: str                 # COBRO or PAGO
    fecha_banco: Optional[date]
    fecha_erp: Optional[date]
    monto: float
    nombre: str
    banco: str
    comprobante: str
    descripcion_banco: str
    detalle_erp: str
    diferencia_dias: int = 0
    alerta: str = ""


# ── ERP file loaders ──────────────────────────────────────────────────────────

def _parse_date(value) -> Optional[date]:
    from datetime import datetime
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _safe_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return abs(float(value))
    if isinstance(value, str):
        try:
            return abs(float(value.strip().replace(",", ".")))
        except ValueError:
            return 0.0
    return 0.0


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_erp_cobros(path: str) -> Tuple[List[ERPRecord], str]:
    """Load COBROS file from ERP Coliseo."""
    return _load_erp_file(path, tipo="COBRO")


def load_erp_pagos(path: str) -> Tuple[List[ERPRecord], str]:
    """Load PAGOS file from ERP Coliseo."""
    return _load_erp_file(path, tipo="PAGO")


def _load_erp_file(path: str, tipo: str) -> Tuple[List[ERPRecord], str]:
    """
    Load a COBROS or PAGOS Excel export from Coliseo ERP.
    Expected columns: Fecha, Comprobante, (Proveedor|Cliente), Nombre,
                      Total, Valor, Detalle Valor, Numero Valor,
                      Fecha Valor, Importe Valor
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return [], f"Error loading {path}: {e}"

    if not rows:
        return [], "Empty file"

    # Find header row
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(str(v).strip() == "Fecha" for v in row if v):
            header_row = {str(v).strip(): j for j, v in enumerate(row) if v}
            header_idx = i
            break

    if header_row is None:
        return [], "Header row not found"

    records: List[ERPRecord] = []
    for row in rows[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue
        try:
            fecha        = _parse_date(row[header_row.get("Fecha", 0)])
            comprobante  = _safe_str(row[header_row.get("Comprobante", 1)])
            nombre       = _safe_str(row[header_row.get("Nombre", 3)])
            total        = _safe_float(row[header_row.get("Total", 4)])
            medio_pago   = _safe_str(row[header_row.get("Valor", 5)])
            detalle      = _safe_str(row[header_row.get("Detalle Valor", 6)])
            fecha_valor  = _parse_date(row[header_row.get("Fecha Valor", 8)])
            importe      = _safe_float(row[header_row.get("Importe Valor", 9)])

            if fecha is None and importe == 0:
                continue

            records.append(ERPRecord(
                fecha         = fecha,
                comprobante   = comprobante,
                nombre        = nombre,
                total         = total,
                medio_pago    = medio_pago,
                detalle_valor = detalle,
                importe_valor = importe,
                fecha_valor   = fecha_valor,
                tipo          = tipo,
            ))
        except Exception:
            continue

    logger.info(f"ERP {tipo}: {len(records)} records loaded from {Path(path).name}")
    return records, "OK"


# ── Bank hint matcher ─────────────────────────────────────────────────────────

def _matches_bank(detalle_valor: str, bank_name: str) -> bool:
    """Check if an ERP detalle_valor string refers to a given bank."""
    dv = detalle_valor.upper()
    for bank_key, hints in BANK_HINTS.items():
        if bank_key.upper() in bank_name.upper():
            return any(h.upper() in dv for h in hints)
    return False


def _date_diff(d1: Optional[date], d2: Optional[date]) -> int:
    """Absolute difference in days between two dates."""
    if d1 is None or d2 is None:
        return 999
    return abs((d1 - d2).days)


# ── Reconciliation engine ─────────────────────────────────────────────────────

class Reconciler:
    """
    Crosses bank transactions against ERP COBROS + PAGOS records.
    Produces a list of ReconciliationLine with status for each item.
    """

    # Amount tolerance for matching (Argentine banking rounding)
    AMOUNT_TOLERANCE = 1.0

    # Date tolerance in days (1 = allows ±1 day, covers MP same-day issue)
    DATE_TOLERANCE = 1

    def reconcile(
        self,
        bank_transactions: list,   # List[Transaction] from Step 1
        erp_cobros: List[ERPRecord],
        erp_pagos: List[ERPRecord],
    ) -> List[ReconciliationLine]:
        """
        Main reconciliation method.
        Returns list of ReconciliationLine sorted by status then date.
        """
        lines: List[ReconciliationLine] = []

        # Work on copies so we can mark records as matched
        cobros = [r for r in erp_cobros]
        pagos  = [r for r in erp_pagos]

        # ── Step A: match bank transactions against ERP ───────────────────────
        for tx in bank_transactions:
            if tx.tipo_movimiento == "INTERNO":
                continue  # skip internal transfers

            # Determine which ERP list to search
            erp_pool = cobros if tx.tipo_movimiento == "COBRO" else pagos
            amount   = abs(tx.importe_neto)

            if amount < 0.01:
                continue

            # Allow ±1 day for Mercado Pago
            max_days = self.DATE_TOLERANCE if tx.pestaña in MP_BANKS else self.DATE_TOLERANCE

            best_match = self._find_best_match(tx, erp_pool, amount, max_days)

            if best_match:
                best_match.matched = True
                diff = _date_diff(tx.fecha, best_match.fecha)
                lines.append(ReconciliationLine(
                    estado            = "CONCILIADO",
                    tipo              = tx.tipo_movimiento,
                    fecha_banco       = tx.fecha,
                    fecha_erp         = best_match.fecha,
                    monto             = amount,
                    nombre            = best_match.nombre,
                    banco             = tx.banco,
                    comprobante       = best_match.comprobante,
                    descripcion_banco = tx.descripcion,
                    detalle_erp       = best_match.detalle_valor,
                    diferencia_dias   = diff,
                    alerta            = f"⚠ Diferencia de {diff} día(s)" if diff > 0 else "",
                ))
            else:
                # Bank has it, ERP doesn't
                lines.append(ReconciliationLine(
                    estado            = "PENDIENTE ERP",
                    tipo              = tx.tipo_movimiento,
                    fecha_banco       = tx.fecha,
                    fecha_erp         = None,
                    monto             = amount,
                    nombre            = "",
                    banco             = tx.banco,
                    comprobante       = "",
                    descripcion_banco = tx.descripcion,
                    detalle_erp       = "",
                    alerta            = "⚠ En banco pero no en Coliseo ERP",
                ))

        # ── Step B: ERP records with no bank match ────────────────────────────
        for erp_record in cobros + pagos:
            if erp_record.matched:
                continue
            if erp_record.importe_valor < 0.01:
                continue
            lines.append(ReconciliationLine(
                estado            = "PENDIENTE BANCO",
                tipo              = erp_record.tipo,
                fecha_banco       = None,
                fecha_erp         = erp_record.fecha,
                monto             = erp_record.importe_valor,
                nombre            = erp_record.nombre,
                banco             = erp_record.detalle_valor,
                comprobante       = erp_record.comprobante,
                descripcion_banco = "",
                detalle_erp       = erp_record.detalle_valor,
                alerta            = "⚠ En Coliseo ERP pero no en banco",
            ))

        # Sort: CONCILIADO first, then PENDIENTE BANCO, then PENDIENTE ERP
        order = {"CONCILIADO": 0, "PENDIENTE BANCO": 1, "PENDIENTE ERP": 2}
        lines.sort(key=lambda l: (order.get(l.estado, 9),
                                  l.fecha_banco or l.fecha_erp or date.min))

        logger.info(
            f"Reconciliation complete: "
            f"{sum(1 for l in lines if l.estado == 'CONCILIADO')} conciliados, "
            f"{sum(1 for l in lines if l.estado == 'PENDIENTE BANCO')} pendiente banco, "
            f"{sum(1 for l in lines if l.estado == 'PENDIENTE ERP')} pendiente ERP"
        )
        return lines

    def _find_best_match(
        self,
        tx,
        erp_pool: List[ERPRecord],
        amount: float,
        max_days: int,
    ) -> Optional[ERPRecord]:
        """
        Find the best matching ERP record for a bank transaction.
        Matching criteria (in priority order):
          1. Amount within tolerance AND date within max_days AND bank hint matches
          2. Amount within tolerance AND date within max_days
          3. Amount within tolerance only (loose match, flagged)
        """
        candidates = []
        for rec in erp_pool:
            if rec.matched:
                continue
            amount_match = abs(rec.importe_valor - amount) <= self.AMOUNT_TOLERANCE
            if not amount_match:
                continue
            date_diff = _date_diff(tx.fecha, rec.fecha_valor or rec.fecha)
            bank_match = _matches_bank(rec.detalle_valor, tx.banco)
            candidates.append((rec, date_diff, bank_match))

        if not candidates:
            return None

        # Priority 1: amount + date + bank all match
        perfect = [(r, d, b) for r, d, b in candidates if d <= max_days and b]
        if perfect:
            return min(perfect, key=lambda x: x[1])[0]

        # Priority 2: amount + date match
        date_ok = [(r, d, b) for r, d, b in candidates if d <= max_days]
        if date_ok:
            return min(date_ok, key=lambda x: x[1])[0]

        # Priority 3: amount only (loose — may be a future-dated item)
        loose = [(r, d, b) for r, d, b in candidates if d <= 5]
        if loose:
            return min(loose, key=lambda x: x[1])[0]

        return None
