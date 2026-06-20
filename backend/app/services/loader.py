# app/services/loader.py
# Excel file loading service for the Delfabro client.
# Reads the 3 input files and makes them available to the rest of the pipeline.

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl import load_workbook

from app.config import SHEET_CONFIG, NON_BANK_SHEETS

logger = logging.getLogger("argus.loader")


class ExcelLoader:
    """Loads and validates the three Excel files for the Delfabro process."""

    def __init__(self):
        self.wb_movimientos: Optional[openpyxl.Workbook] = None
        self.wb_saldos: Optional[openpyxl.Workbook] = None
        self.wb_caja: Optional[openpyxl.Workbook] = None
        self._paths: Dict[str, Path] = {}

    # ── Loading ──────────────────────────────────────────────────────────────

    def load_movimientos(self, path: str) -> Tuple[bool, str]:
        """Load the Daily Bank Transactions file."""
        try:
            p = Path(path)
            if not p.exists():
                return False, f"Archivo no encontrado: {path}"
            self.wb_movimientos = load_workbook(p, read_only=True, data_only=True)
            self._paths["movimientos"] = p
            detected = self._detect_bank_sheets()
            logger.info(f"Movimientos loaded: {p.name} — {len(detected)} bank sheets")
            return True, f"OK — {len(detected)} cuentas detectadas"
        except Exception as e:
            logger.error(f"Error loading movimientos: {e}")
            return False, f"Error: {e}"

    def load_saldos(self, path: str) -> Tuple[bool, str]:
        """Load the Daily Balances file."""
        try:
            p = Path(path)
            if not p.exists():
                return False, f"Archivo no encontrado: {path}"
            self.wb_saldos = load_workbook(p, read_only=True, data_only=True)
            self._paths["saldos"] = p
            logger.info(f"Saldos loaded: {p.name}")
            return True, "OK"
        except Exception as e:
            logger.error(f"Error loading saldos: {e}")
            return False, f"Error: {e}"

    def load_caja(self, path: str) -> Tuple[bool, str]:
        """Load the Caja Fábrica Digital file."""
        try:
            p = Path(path)
            if not p.exists():
                return False, f"Archivo no encontrado: {path}"
            self.wb_caja = load_workbook(p, read_only=True, data_only=True)
            self._paths["caja"] = p
            logger.info(f"Caja loaded: {p.name}")
            return True, "OK"
        except Exception as e:
            logger.error(f"Error loading caja: {e}")
            return False, f"Error: {e}"

    # ── Sheet detection ───────────────────────────────────────────────────────

    def _detect_bank_sheets(self) -> List[str]:
        """Return the known bank sheets that exist in the workbook."""
        if not self.wb_movimientos:
            return []
        available = set(self.wb_movimientos.sheetnames)
        known = set(SHEET_CONFIG.keys())
        found = [s for s in self.wb_movimientos.sheetnames if s in known]
        unknown_new = available - known - NON_BANK_SHEETS
        if unknown_new:
            logger.warning(f"Unknown sheets (will not be processed): {unknown_new}")
        return found

    def get_bank_sheets(self) -> List[str]:
        """List of detected bank sheets."""
        return self._detect_bank_sheets()

    # ── Row reading ──────────────────────────────────────────────────────────

    def get_sheet_rows(self, sheet_name: str) -> List[tuple]:
        """Return all rows from a sheet in the movimientos workbook."""
        if not self.wb_movimientos:
            return []
        if sheet_name not in self.wb_movimientos.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return []
        ws = self.wb_movimientos[sheet_name]
        return list(ws.iter_rows(values_only=True))

    def get_categories(self) -> Dict[int, str]:
        """Read the accounting categories table and return {code: name}."""
        cats: Dict[int, str] = {}
        if not self.wb_movimientos:
            return cats
        sheet_name = "CATEGORIAS CONTABLES"
        if sheet_name not in self.wb_movimientos.sheetnames:
            logger.warning("Sheet CATEGORIAS CONTABLES not found")
            return cats
        ws = self.wb_movimientos[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                try:
                    code = int(row[0])
                    name = str(row[1]).strip()
                    cats[code] = name
                except (ValueError, TypeError):
                    continue
        logger.info(f"Categories loaded: {len(cats)}")
        return cats

    def get_saldos_rows(self) -> List[tuple]:
        """Return rows from the BANCOS DEL DIA sheet."""
        if not self.wb_saldos:
            return []
        sheet_name = "BANCOS DEL DIA"
        if sheet_name not in self.wb_saldos.sheetnames:
            logger.warning("Sheet BANCOS DEL DIA not found")
            return []
        ws = self.wb_saldos[sheet_name]
        return list(ws.iter_rows(values_only=True))

    def get_caja_months(self) -> List[str]:
        """Return the available month sheets in the Caja workbook."""
        if not self.wb_caja:
            return []
        return [s for s in self.wb_caja.sheetnames if s not in {"acumulado", "Categorías"}]

    def is_ready(self) -> bool:
        """True if all 3 files are loaded."""
        return all([self.wb_movimientos, self.wb_saldos, self.wb_caja])
